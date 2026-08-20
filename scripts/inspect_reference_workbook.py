from __future__ import annotations

import json
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


REFERENCE_PATH = Path("/home/ubuntu/upload/01-btl_mort_rates.xlsx")
if not REFERENCE_PATH.exists():
    REFERENCE_PATH = Path("/home/ubuntu/upload/01-btl-mort_rates.xlsx")
OUTPUT_PATH = Path("/home/ubuntu/lender-data-extractor/docs/reference-workbook-profile.json")


def cell_style(cell):
    return {
        "number_format": cell.number_format,
        "font": {
            "name": cell.font.name,
            "size": cell.font.sz,
            "bold": cell.font.bold,
            "italic": cell.font.italic,
            "color": getattr(cell.font.color, "rgb", None) if cell.font.color else None,
        },
        "fill": {
            "type": cell.fill.fill_type,
            "fgColor": cell.fill.fgColor.rgb,
            "bgColor": cell.fill.bgColor.rgb,
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal,
            "vertical": cell.alignment.vertical,
            "wrap_text": cell.alignment.wrap_text,
        },
        "border": str(cell.border),
    }


def worksheet_profile(ws):
    rows = []
    non_empty_rows = []
    for row in ws.iter_rows():
        values = [cell.value for cell in row]
        if any(value not in (None, "") for value in values):
            row_number = row[0].row
            non_empty_rows.append(row_number)
            rows.append({
                "row": row_number,
                "values": values,
                "styles": [cell_style(cell) for cell in row],
            })

    headers = []
    for record in rows[:10]:
        non_empty = [value for value in record["values"] if value not in (None, "")]
        if len(non_empty) >= 3:
            headers.append({"row": record["row"], "values": record["values"]})

    return {
        "title": ws.title,
        "dimensions": {"rows": ws.max_row, "columns": ws.max_column},
        "freeze_panes": ws.freeze_panes,
        "merged_cells": sorted(str(rng) for rng in ws.merged_cells.ranges),
        "sheet_properties": {
            "tabColor": ws.sheet_properties.tabColor.rgb if ws.sheet_properties.tabColor else None,
            "page_setup": str(ws.page_setup),
        },
        "column_widths": {
            key: dimension.width
            for key, dimension in ws.column_dimensions.items()
            if dimension.width is not None
        },
        "row_heights": {
            str(key): dimension.height
            for key, dimension in ws.row_dimensions.items()
            if dimension.height is not None
        },
        "candidate_header_rows": headers,
        "non_empty_rows": non_empty_rows,
        "sample_non_empty_rows": rows[:20],
        "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter.ref else None,
        "data_validations": [str(validation.sqref) for validation in ws.data_validations.dataValidation],
    }


def main():
    if not REFERENCE_PATH.exists():
        raise FileNotFoundError(f"Reference workbook not found: {REFERENCE_PATH}")
    wb = load_workbook(REFERENCE_PATH, data_only=False)
    profile = {
        "workbook": str(REFERENCE_PATH),
        "sheet_names": wb.sheetnames,
        "worksheets": [worksheet_profile(ws) for ws in wb.worksheets],
    }
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(profile, indent=2, default=str), encoding="utf-8")
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()

