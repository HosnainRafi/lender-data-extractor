from pathlib import Path

from openpyxl import load_workbook


REFERENCE_PATH = Path("/home/ubuntu/upload/01-btl-mort_rates.xlsx")
if not REFERENCE_PATH.exists():
    REFERENCE_PATH = Path("/home/ubuntu/upload/01-btl-mort_rates.xlsx")


def row_values(ws, row_number):
    return [ws.cell(row=row_number, column=column).value for column in range(1, ws.max_column + 1)]


def main():
    workbook = load_workbook(REFERENCE_PATH, data_only=False)
    for sheet_name in ["Current Products", "New Products", "Withdrawn Products", "Additional"]:
        ws = workbook[sheet_name]
        print(f"SHEET: {sheet_name}")
        print(f"DIMENSIONS: {ws.max_row} rows x {ws.max_column} columns")
        for row_number in range(1, min(ws.max_row, 40) + 1):
            values = row_values(ws, row_number)
            if any(value not in (None, "") for value in values):
                print(f"ROW {row_number}: {values}")
        print(f"FREEZE: {ws.freeze_panes}")
        print(f"FILTER: {ws.auto_filter.ref}")
        print(f"MERGES: {[str(rng) for rng in ws.merged_cells.ranges]}")
        print("---")


if __name__ == "__main__":
    main()
